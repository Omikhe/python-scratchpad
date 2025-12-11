import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}


for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    invetory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    invetory_price = product_list.cell(product_row, 5)

    # calculating number of products per supplier
    if supplier_name in products_per_supplier:
        cuurent_num_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = cuurent_num_products + 1
    else:
        print("adding a new supplier")
        products_per_supplier[supplier_name] = 1
    
    # calculating total invetory value per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] =  current_total_value + invetory * price
    else:
        total_value_per_supplier[supplier_name] = invetory * price
    
    # products with inventory less than 10
    if invetory < 10:
        products_under_10_inv[int(product_num)] = int(invetory)
    
    # add value for total invetory price
    invetory_price.value = invetory * price

print(" ")
print(products_per_supplier)
print(total_value_per_supplier)
print(products_under_10_inv)

inv_file.save("invetory_with_total_value.xlsx")